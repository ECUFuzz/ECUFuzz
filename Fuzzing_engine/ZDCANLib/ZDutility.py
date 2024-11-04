import time
import can
from threading import Thread
import json
import os, sys
import xlwings as xw
import openpyxl
import random
import ctypes
# from flask import Flask, render_template

SecDllPath = './ZDCANLib/GenerateKeyExImpl.dll'

class zd_utility:
    @staticmethod
    def SetDllPath(path):
        global SecDllPath
        SecDllPath = path

    @staticmethod
    def ConvertBytes2Hex(payload):
        strval = ""
        if payload is not None:
            for element in payload:
                strval += "{:0>2X}".format(element)
                strval += " "
        return strval.strip()

    @staticmethod
    def ConvertBytes2Ascii(payload):
        strval = ""
        if payload is not None:
            for element in payload:
                strval += chr(element)
        return strval
    
    @staticmethod
    def ConvertAscii2Bytes(payload):
        if type(payload) is str:
            bytelist = []
            for i in range(0, len(payload)):
                bytelist.append(ord(payload[i:i+1]))
            return bytelist
        else:
            return payload
        
    @staticmethod
    def ConvertStr2Bytes(strval):
        strval=strval.replace('\r','')
        strval=strval.replace('\n','')
        strval=strval.replace('0x','')
        strval=strval.replace(',','')
        strval=strval.replace(' ','')
        cmdlist=[]
        for i in range(0,(len(strval) - 1),2):
            cmdlist.append(int(strval[i:i+2], 16))
        return cmdlist
    
    def LoadJson(jsonfile):
        if(os.path.exists(jsonfile)):
            return json.load(open(jsonfile))
        else:
            return None
    
    def SaveJson(directory, file):
        str = json.dumps(directory, indent=2)
        with open(file, 'w') as f:
            f.write(str)

    @staticmethod
    def SecAccessCrc32(input):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        length = len(input)
        uint8_arr = ctypes.c_uint8 * length
        input_arr = uint8_arr()
        for i in range(0, length):
            input_arr[i] = input[i]
        return dll.SecAccessCrc32(ctypes.byref(input_arr), length)

    @staticmethod
    def Crc8(input):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        length = len(input)
        uint8_arr = ctypes.c_uint8 * length
        input_arr = uint8_arr()
        for i in range(0, length):
            input_arr[i] = input[i]
        crc8 = ctypes.c_uint8(dll.Crc8(ctypes.byref(input_arr), length, 0xD5, 0xFF))
        return crc8.value        

    @staticmethod
    def Crc32(input):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        length = len(input)
        uint8_arr = ctypes.c_uint8 * length
        input_arr = uint8_arr()
        for i in range(0, length):
            input_arr[i] = input[i]
        crc32 = ctypes.c_uint32(dll.Crc32(ctypes.byref(input_arr), length, 0x4c11db7, 0xFFFFFFFF))
        return crc32.value
    
    @staticmethod
    def S37Crc32(input):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        length = len(input)
        uint8_arr = ctypes.c_uint8 * length
        input_arr = uint8_arr()
        for i in range(0, length):
            input_arr[i] = input[i]
        crc32 = ctypes.c_uint32(dll.S37Crc32(ctypes.byref(input_arr), length))
        return crc32.value

    @staticmethod
    def SecAccessGenerateZDKey(seed):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        uint8_seed_arr = ctypes.c_uint8 * 4
        seed_arr = uint8_seed_arr(seed[0], seed[1], seed[2], seed[3])
        uint8_key_arr = ctypes.c_uint8 * 4
        key_arr = uint8_key_arr(0,0,0,0)
        uint32_key_length = ctypes.c_uint32
        key_length = uint32_key_length(0)
        dll.GenerateZDKey(ctypes.byref(seed_arr), 4, ctypes.byref(key_arr), ctypes.byref(key_length))
        key = []
        for i in range(0, key_length.value):
            key.append(key_arr[i])
        return key

    @staticmethod
    def SecAccessGenerateBootKey(seed):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        uint8_seed_arr = ctypes.c_uint8 * 2
        seed_arr = uint8_seed_arr(seed[0], seed[1])
        uint8_key_arr = ctypes.c_uint8 * 2
        key_arr = uint8_key_arr(0,0)
        uint32_key_length = ctypes.c_uint32
        key_length = uint32_key_length(0)
        dll.GenerateBootKey(ctypes.byref(seed_arr), 2, ctypes.byref(key_arr), ctypes.byref(key_length))
        key = []
        for i in range(0, key_length.value):
            key.append(key_arr[i])
        return key

    # @staticmethod
    # def SecAccessGenerateCmacKey(seed):
    #     dll = ctypes.cdll.LoadLibrary(SecDllPath)
    #     uint8_seed_arr = ctypes.c_uint8 * 16
    #     seed_arr = uint8_seed_arr(seed[0], seed[1], seed[2], seed[3], seed[4], seed[5], seed[6], seed[7], seed[8], seed[9], seed[10], seed[11], seed[12], seed[13], seed[14], seed[15])
    #     uint8_key_arr = ctypes.c_uint8 * 16
    #     key_arr = uint8_key_arr(0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0)
    #     uint32_key_length = ctypes.c_uint32
    #     key_length = uint32_key_length(0)
    #     dll.GenerateSecKey(ctypes.byref(seed_arr), 16, ctypes.byref(key_arr), ctypes.byref(key_length))
    #     key = []
    #     for i in range(0, key_length.value):
    #         key.append(key_arr[i])
    #     return key

    @staticmethod
    def GenerateSecKey(type, seed):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        uint8_seed_arr = ctypes.c_uint8 * len(seed)
        seed_arr = uint8_seed_arr()
        uint8_key_arr = ctypes.c_uint8 * len(seed)
        key_arr = uint8_key_arr()
        for i in range(len(seed)):
            seed_arr[i] = seed[i]
            key_arr[i] = 0
        uint32_key_length = ctypes.c_uint32
        key_length = uint32_key_length(0)
        dll.GenerateSecKey(type, ctypes.byref(seed_arr), len(seed), ctypes.byref(key_arr), ctypes.byref(key_length))
        key = []
        for i in range(0, key_length.value):
            key.append(key_arr[i])
        return key


    @staticmethod
    def GenerateRsaSign(type, data):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        uint8_data_arr = ctypes.c_uint8 * len(data)
        data_arr = uint8_data_arr()
        for i in range(len(data)):
            data_arr[i] = data[i]

        uint8_sign_arr = ctypes.c_uint8 * 4096
        sign_arr = uint8_sign_arr()
        for i in range(4096):
            sign_arr[i] = 0

        uint32_sign_length = ctypes.c_uint32
        sign_length = uint32_sign_length(0)

        type = type.encode('utf-8')
        ret = dll.GenerateRsaSign(type, ctypes.byref(data_arr), len(data), ctypes.byref(sign_arr), ctypes.byref(sign_length))
        rsa_sign = []
        for i in range(0, sign_length.value):
            rsa_sign.append(sign_arr[i])
        return rsa_sign

    @staticmethod
    def GenerateRsaSignWAddr(type, startaddress, data):
        datalen = len(data)
        datalist = [(startaddress & 0xFF000000) >> 24, (startaddress & 0x00FF0000) >> 16, (startaddress & 0x0000FF00) >> 8, (startaddress & 0x000000FF) >> 0, 
                    (datalen & 0xFF000000) >> 24, (datalen & 0x00FF0000) >> 16, (datalen & 0x0000FF00) >> 8, (datalen & 0x000000FF) >> 0] + data
        return zd_utility.GenerateRsaSign(type, datalist)

    @staticmethod
    def HighGAppendCrc3(data):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        crc3 = dll.HighGCrc3(data & 0xFFFFFFF8)
        return (data & 0xFFFFFFF8) | crc3
    
    @staticmethod
    def Crc3(data, mask0, mask1, mask2, xor):
        dll = ctypes.cdll.LoadLibrary(SecDllPath)
        return (dll.Crc3((data & 0xFFFFFFFF), (mask0 & 0xFFFFFFFF), (mask1 & 0xFFFFFFFF), (mask2 & 0xFFFFFFFF), (xor & 0xFF)))

    @staticmethod
    def process_bar(step):
        for i in range(0, 101, step):
            print("\r", end="")
            print("work progress: {}%: ".format(i), "▋" * (i // step), end="")
            sys.stdout.flush()
            time.sleep(0.1)
        print('\n')
    
    @staticmethod
    def DecodeS37(iostream=None, string=None):
        startaddress = None
        transferdata = []
        if iostream is not None:
            lines = iostream.readlines()
        elif string is not None:
            lines = string.split('\n')
        for line in lines:
            s37LineHead = line[0:4]
            if type(s37LineHead) is str:
                expLineHead = 'S325'
            elif type(s37LineHead) is bytes:
                expLineHead = b'S325'
            else:
                expLineHead = None
            if s37LineHead == expLineHead:
                if startaddress is None:
                    startaddress = int(line[4:12], 16)
                for i in range(0, 64, 2):
                    transferdata.append(int(line[12+i:12+i+2], 16))
        return startaddress, transferdata
    
    @staticmethod
    def DecodeS37StreamList(iostreamlist):
        startaddress = None
        transferdata = []
        for iostream in iostreamlist:
            startaddress_, transferdata_ = zd_utility.DecodeS37(iostream)
            if startaddress is None:
                startaddress = startaddress_
            transferdata += transferdata_
        return startaddress, transferdata
    
    @staticmethod
    def DecodeS37FileList(filelist):
        startaddress = None
        transferdata = []
        for file in filelist:
            if(file.endswith('.s37')):
                with open(file, 'r') as f:
                    startaddress_, transferdata_ = zd_utility.DecodeS37(f)
                    if startaddress is None:
                        startaddress = startaddress_
                    transferdata += transferdata_
        return startaddress, transferdata


    @staticmethod
    def MergeS37(iostreamList):
        linestart = 'S00F0000415343454E544153434505070A\n'
        lineend = 'S70500000000FA\n'
        addrList = []
        s37List = []
        index = 0
        for file in iostreamList:
            lines = file.split('\n')
            startaddr = None
            s37str = ''
            for line in lines:
                if line[0:4] == 'S325':
                    s37str += line + '\n'
                    if startaddr is None:
                        startaddr = int(line[4:12], 16)
            if len(addrList) == 0:
                addrList.append(startaddr)
                s37List.append(s37str)
            elif len(addrList) == 1:
                if startaddr > addrList[0]:
                    addrList.append(startaddr)
                    s37List.append(s37str)
                else:
                    addrList.insert(0, startaddr)
                    s37List.insert(0, s37str)
            else:
                if startaddr < addrList[0]:
                    addrList.insert(0, startaddr)
                    s37List.insert(0, s37str)
                elif startaddr > addrList[len(addrList)-1]:
                    addrList.append(startaddr)
                    s37List.append(s37str)
                else:
                    for i in range(len(addrList) - 1):
                        if (addrList[i] < startaddr) and (startaddr < addrList[i+1]):
                            addrList.insert(i+1, startaddr)
                            s37List.insert(i+1, s37str)
            # index += 1
            # for addr in addrList:
            #     print(hex(addr))

        strall = ''
        strall += linestart
        for s37 in s37List:
            strall += s37
        strall += lineend
        return strall    

    @staticmethod
    def MergeS37StreamList(streamlist):
        iostreamList = []
        for s37 in streamlist:
            iostreamList.append(s37.read().decode('utf-8'))
        return zd_utility.MergeS37(iostreamList)
    
    @staticmethod
    def MergeS37FileList(filelist):
        iostreamList = []
        for s37 in filelist:
            with open(s37, 'r') as f:
                iostreamList.append(f.read())
        return zd_utility.MergeS37(iostreamList)

class TimeoutTimer():
    def __init__(self, timeoutTime=1):
        self.timeoutTime = timeoutTime

    def Start(self, timeoutTime=None):
        if(timeoutTime is not None):
            self.timeoutTime = timeoutTime
        self.startTime = time.time()
        # print('time restart ' + str(self.startTime) + ' timeout ' + str(self.timeoutTime))
    
    def Restart(self, timeoutTime=None):
        self.Start(timeoutTime)
    
    def IsExpired(self):
        currTime = time.time()
        if(currTime - self.startTime) > self.timeoutTime:
            # print('time expired ' + str(currTime) + ' ' + str(self.startTime) + ' timeout ' + str(self.timeoutTime))
            return True
        else:
            # print('time expired ' + str(currTime) + ' ' + str(self.startTime) + ' timeout ' + str(self.timeoutTime))
            return False
        
    def IsNotExpired(self):
        if self.IsExpired():
            # raise Exception("Timeout")
            return False
        else:
            return True

class UdsSimulation():
    def __init__(self, interface='virtual', channel='virtual_ch', txId=0x760, rxId=0x660):
        self.bus = can.Bus(interface=interface, app_name='CANalyzer', channel=channel, bitrate=500000, sjw_abr=1, tseg1_abr=16, tseg2_abr=3, sjw_dbr=1, tseg1_dbr=16, tseg2_dbr=3)
        self.txId = txId
        self.rxId = rxId
        self.t1 = Thread(target=self.Response)
        self.ThreadRun = True

    def Start(self):
        self.ThreadRun = True
        self.t1.start()

    def Stop(self):
        self.ThreadRun = False
        self.t1.join()

    def Response(self):
        while self.ThreadRun:
            rx_msg = self.bus.recv(10)
            # print(rx_msg)
            if(rx_msg is None):
                break
            if(rx_msg.arbitration_id == self.rxId):
                # print(rx_msg)
                framedata = [0x03, 0x7F, 0x10, 0x7F, 0x00, 0x00, 0x00, 0x00]
                if(rx_msg.data[0] & 0xF0) == 0x10: # First frame
                    framedata = [0x30, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00]
                elif (rx_msg.data[0] & 0xF0) == 0x30: # Flow control frame
                    framedata = [0x21, 0x08, 0x14, 0x06, 0x08, 0x0F, 0x09, 0x00]
                elif ((rx_msg.data[0] & 0xF0) == 0x20): # Continuous frame
                    if (self.targetSN == self.SN):
                        self.SN = 0
                        framedata = [self.srvID + 0x40, self.subFunc0, self.subFunc1, 0x00, 0x00, 0x00, 0x00, 0x00]
                    else:
                        self.SN += 1
                elif ((rx_msg.data[0] & 0xF0) == 0x00): # single frame
                    self.srvID = rx_msg.data[1]
                    self.subFunc0 = rx_msg.data[2]
                    self.subFunc1 = rx_msg.data[3]
                    if self.srvID == 0x22:
                        if(self.subFunc0 == 0xFD) and (self.subFunc1 == 0x39):
                            framedata = [0x10, 0x0C, 0x62, 0xFD, 0x39, 0x08, 0x13, 0x03]
                    elif self.srvID == 0x19:
                        if(self.subFunc0 == 0x02) and (self.subFunc1 == 0xCB):
                            framedata = [0x07, 0x59, 0x02, 0xCB, 0x80, 0x95, 0x13, 0x01]
                            # framedata = [0x07, 0x59, 0x02, 0xCB, 0x92, 0x04, 0x11, 0x08]
                    elif self.srvID == 0x27:
                        if(self.subFunc0 == 0x41):
                            framedata = [0x06, 0x67, 0x41, 0x11, 0x22, 0x33, 0x44, 0x40]
                        elif (self.subFunc0 == 0x42):
                            framedata = [0x02, 0x67, 0x42, 0x11, 0x22, 0x33, 0x44, 0x40]
                msg = can.Message(arbitration_id=self.txId, is_extended_id=False, data=framedata)
                self.bus.send(msg)
        time.sleep(1)
        self.bus.shutdown()

class zd_excel():

    class zd_sheet():
        def __init__(self, excelapp, ws):
            self.ws = ws
            self.excelapp = excelapp

        def GetMaxRow(self):
            if self.excelapp == 'xlwings':
                return self.ws.used_range.last_cell.row
            elif self.excelapp == 'openpyxl':
                return self.ws.max_row

        def GetMaxColumn(self):
            if self.excelapp == 'xlwings':
                return self.ws.used_range.last_cell.column
            elif self.excelapp == 'openpyxl':
                return self.ws.max_column
        
        def GetCellText(self, row, col):
            if self.excelapp == 'xlwings':
                return self.ws.range(row, col).options(numbers=lambda x: str(int(x))).value
            elif self.excelapp == 'openpyxl':
                return str(self.ws.cell(row, col).value)

    def __init__(self, filepath, excelapp='openpyxl'):
        self.excelapp = excelapp
        if self.excelapp == 'xlwings':
            app = xw.App(visible=False, add_book=False)
            app.screen_updating = False
            app.display_alerts = False
            self.wb = app.books.open(filepath)
        elif self.excelapp == 'openpyxl':
            self.wb = openpyxl.load_workbook(filepath)
    
    def GetSheet(self, sheetname):
        if self.excelapp == 'xlwings':
            return self.zd_sheet(self.excelapp, self.wb.sheets[sheetname])
        elif self.excelapp == 'openpyxl':
            return self.zd_sheet(self.excelapp, self.wb[sheetname])
    
    def Close(self):
        self.wb.close()  

class HtmlLiveReport():
    def __init__(self, file_path, flashcycle=None):
        self.file_path = file_path
        self.isFirstTable = False
        self.data = ''
        self.data += "<!DOCTYPE html>\n"
        self.data += "<html>\n"
        if(flashcycle != None):
            self.data += "<head>\n"
            self.data += '<meta http-equiv="refresh" content="'+str(flashcycle)+'">'
            self.data += "</head>\n"
        self.data += "<body>\n"

    def GenerateHtml(self):
        self.data += '</tbody>\n'
        self.data += '</table>\n'
        self.data += "</body>\n"
        self.data += "</html>\n"
        self.html_f = open(self.file_path, 'w')
        self.html_f.write(self.data)
        self.html_f.close()

    # Output the table header. 
    # If this is not the first table, it means there are previous tables. 
    # Therefore, first, complete the ending of the previous table.
    def PrintTableHead(self, table_name, cellsize='250:250:250:250'):
        if self.isFirstTable == False:
            self.isFirstTable = True
        else:
            self.data += '</tbody>\n'
            self.data += '</table>\n'
        self.number = 0
        self.data += '<table style="border-collapse: collapse; width: 1500px;" border="1">\n'
        self.data += '<tbody>\n'
        self.data += '<tr>\n'
        self.data += '<td style="width: 1500px; background-color: #bfedd2; border-style: solid;" colspan="1500"><strong><span style="font-size: 14pt;">' + table_name + '</span></strong></td>\n'
        self.data += '</tr>\n'
        self.cellsizelist = cellsize.split(':')
    
    def PrintTableCell(self, datalist, color):
        self.number += 1
        self.data += '<tr>\n'
        self.data += '<td style="width: ' + self.cellsizelist[0] + 'px; background-color: ' + color + '; border-style: solid;"colspan="' + self.cellsizelist[0] + '"><span style="font-size: 10pt;">' + str(self.number) + '</span></td>\n'
        if((len(self.cellsizelist) - 1) == len(datalist)):
            for i in range(1, len(self.cellsizelist)):
                self.data += '<td style="width: ' + self.cellsizelist[i] + 'px; background-color: ' + color + '; border-style: solid;"colspan="' + self.cellsizelist[i] + '"><span style="font-size: 10pt;">' + datalist[i - 1] + '</span></td>\n'
        else:
            raise Exception("the column number of table is not matched with content list number")
        self.data += '</tr>\n'

    def PrintTableTestReport(self, stepDescription, expectVal, actualVal, log=' '):
        if(self.number == 0):
            self.PrintTableCell(['step description', 'expect value', 'actual value', 'test result', 'log'], 'white')         
        result = ''
        ret = 0
        if expectVal == actualVal:
            color = '#98FB98'
            result = 'pass'
            ret = 0
        else:
            color = 'red'
            result = 'fail'
            ret = 1
        self.PrintTableCell([stepDescription, expectVal, actualVal, result, log], color)
        return ret
    
    def PrintTableUdsReport(self, stepDescription, expectVal, udsResponse):
        logArray = udsResponse[1]
        log = ''
        strActual = ''
        if logArray is not None:
            if len(logArray) > 1:
                strActual = logArray[1]                
            for logEle in logArray:
                log += logEle + "<br />"
        if (expectVal == True) or (expectVal == False):
            return self.PrintTableTestReport(stepDescription, str(expectVal), str(udsResponse[0]), log)
        else:
            return self.PrintTableTestReport(stepDescription, expectVal, strActual, log)


if __name__ == "__main__":
    # test2 = "00 00 02 80 FF 7F 02 80 44 82 02 80 32 00 43 48 59 5F 54 30 30 31 5F 50 31 30 03 00"
    # print("{:0>8X}".format(zd_utility.S37Crc32(zd_utility.ConvertStr2Bytes(test2))))
    
    # dll.main(b"D:\gitlab\ZDCANApp\VehFixedCal.s37", 0xA0020000, 0xA002001C)
    # if sys.argv[1] == "merges37":
    #     fileList = sys.argv[2].split(',')
    #     iostreamList = []
    #     for file in fileList:
    #         if os.path.exists(file):
    #             with open(file, 'r') as f:
    #                 iostreamList.append(f.read())
    #         else:
    #             print(file + ' is not existed')
    #     with open('merge.s37', 'w') as f:
    #         f.write(zd_utility.MergeS37(iostreamList))

    # publickey = "E1016463E7AAB659D2631BC1DEE36BDAC3E5C6765149B01047A21F4EFF5406B6129EFA01843ACC0C225215661887A35077F6A942D49834B1BBD06D794BBF8491C3FAA93306C9FDD463A9CE9CC8C9506BBF28277F78F60E8A6D5DEFC978D220E587DE20408EA4AC547C0BD1CEA94464B306E407D95EC99DEA0C3FEDBA402AEAE7275B486BA592FE52C9E24CCFAEAA8B984D61D71AB2E7AED329A68A23E4BFCE2A11CC6E5AB835585CC221FECAC3DE892252314F7131782E216CAF9478F080BBE84162B44E39E69FDB84F9D04D17BF95FB0BD7B6121826AB2545457D80742EB32EC60F6396A108125F06D6AD1B10D88F91296A0EC7AD77C95D8B7E7B28FB465964FD08172B067E6D0346E601F162E0FF3C023FE635772D90E09AE750C66D4342C2D13B14F745DA295E96E7D5D219D953780698B6C80BC48ECFCD5B8425BE86FA21F3A4372CAC1272421183FCEF1515E6345E07555BEB5085D531E27E6E1FE65899B82B8DA64FA70E4CD8425D94B5E08DF9A65A1C1A32C33F5948E4EEDFF27D22B87AA6F1F4B65D1EDA65AFDAD3AA20291B11DA976B57BD8798986D554BF13E59F1DF101D5B014B820900F13539F07755503D6C8A6D4E3A591CF118ABEC206EEE4BAA98C05C8C382AB65DB3AEC9647FFF8DEF248C35B1E028CA280BA723DE7856526976C3FCF2543B0D3F5DB1B9AFAAA096D5D1DD5198BEF84F25B29A1F46DB5E2F"
    # strarray = ''
    # for i in range(0, len(publickey), 2):
    #     strarray += '0x' + publickey[i:(i+2)] + ', '
    # print(strarray)

    # publickeyarr = [0xB0, 0xB2, 0xB2, 0xC5, 0x44, 0xBD, 0x1E, 0x85, 0xD2, 0x73, 0x5B, 0x0A, 0xC0, 0x2E, 0x6F, 0x4C, 0xF9, 0xA5, 0xE1, 0x42, 0x45, 0x02, 0xE9, 0xC6, 0xC8, 0x6D, 0x85, 0x83, 0x1B, 0x7E, 0xD1, 0x60, 0xCC, 0xF8, 0x59, 0x02, 0x7E, 0xDD, 0xB0, 0xC1, 0x55, 0x8B, 0x77, 0xEB, 0x74, 0xA9, 0x97, 0x97, 0xFB, 0x09, 0x3F, 0x13, 0xAF, 0x45, 0xAB, 0x47, 0x60, 0x58, 0xAA, 0xC6, 0xCA, 0x5A, 0xCC, 0x71, 0x39, 0xFE, 0x32, 0xA0, 0xBB, 0xC2, 0x56, 0xB8, 0x04, 0x4C, 0xDA, 0x31, 0x83, 0x2C, 0xCE, 0xC1, 0xD5, 0x2A, 0x75, 0xCC, 0x95, 0xC1, 0x25, 0xE8, 0xF6, 0xA2, 0x60, 0x69, 0xD0, 0x5D, 0x01, 0x07, 0x3B, 0xD8, 0xF9, 0x3C, 0xF0, 0xEF, 0x84, 0x5C, 0x09, 0xF2, 0xAF, 0x00, 0xAF, 0x27, 0xAF, 0xF5, 0x97, 0x9B, 0xEF, 0xAF, 0x67, 0xCD, 0x0C, 0x82, 0xE5, 0x0B, 0x56, 0xDE, 0x80, 0xC9, 0x90, 0xAB, 0x7B, 0x18, 0x08, 0x3F, 0x8F, 0x18, 0xCB, 0x1E, 0xD0, 0xAC, 0xA9, 0x3A, 0xBB, 0x93, 0x22, 0x27, 0x69, 0xC8, 0x14, 0x2A, 0xAA, 0xC4, 0xE9, 0x73, 0x48, 0x02, 0x3E, 0xDA, 0x4C, 0x39, 0xE4, 0x08, 0x5E, 0x55, 0xB1, 0xD8, 0x59, 0x54, 0xB4, 0x90, 0xF7, 0x67, 0x10, 0xFF, 0x8A, 0xE8, 0x30, 0xB1, 0x71, 0xA9, 0x94, 0xCE, 0xE7, 0xCD, 0xAB, 0x9C, 0xED, 0xB0, 0xB3, 0x5A, 0x5A, 0x7C, 0x89, 0x92, 0x7E, 0xEE, 0x00, 0x11, 0x90, 0x51, 0x9D, 0xF5, 0x9C, 0x9D, 0x77, 0xF9, 0x32, 0x3C, 0x12, 0x54, 0x18, 0x9A, 0x72, 0xB8, 0x87, 0x35, 0x20, 0x82, 0x89, 0x5A, 0x2E, 0x5B, 0xA8, 0x47, 0x9E, 0x6F, 0x13, 0xAB, 0x22, 0xFE, 0xDC, 0xA9, 0x87, 0xB6, 0xF6, 0xF8, 0x29, 0xF3, 0x5C, 0x97, 0x29, 0xBB, 0x85, 0xF7, 0x66, 0xD6, 0x6B, 0x92, 0xF3, 0xEB, 0x27, 0xF4, 0x7D, 0x42, 0x8B, 0xA6, 0xE8, 0x37, 0xA4, 0xF1, 0x1D, 0xA6, 0x7D, 0xA4, 0x0B, 0x63, 0xC6, 0x7C, 0x98, 0xB7, 0xA0, 0xDD, 0x2E, 0x39, 0x6C, 0x6A, 0x84, 0x95, 0xDD, 0xAB, 0xC1, 0x91, 0x00, 0x3F, 0xF1, 0xF5, 0x18, 0x9D, 0xCB, 0xD9, 0x9E, 0x86, 0xF0, 0x31, 0x48, 0xE2, 0x77, 0x94, 0xDF, 0x4B, 0x08, 0x75, 0xAF, 0x4B, 0x57, 0x6A, 0x51, 0x11, 0x2A, 0x39, 0x8F, 0x5F, 0x05, 0xB7, 0x9F, 0xEB, 0x9A, 0x8C, 0xA7, 0xF1, 0x88, 0x3C, 0xE8, 0xD0, 0x1B, 0x74, 0xE0, 0x32, 0x97, 0x09, 0xE9, 0x02, 0x11, 0x4D, 0xAF, 0xE0, 0xE3, 0x44, 0x32, 0xDF, 0xD6, 0x8B, 0x70, 0x93, 0xE3, 0x2C, 0xF3, 0xBA, 0x4A, 0xC5, 0xDA, 0xFB, 0x7A, 0xB8, 0x6B, 0x14, 0xC5, 0x19, 0xB3, 0xDF, 0xF8, 0x96, 0x2B, 0xA9, 0x0A, 0xC5, 0xC5, 0x7C, 0xF8, 0x90, 0x64, 0x4D, 0xBD, 0x33, 0x2C, 0x54, 0xE2, 0x7D, 0x42, 0x24, 0xF1, 0x1D, 0x38, 0xC6, 0x4A, 0xDB, 0xAB, 0x8E, 0xC5, 0x94, 0xAB, 0xB2, 0x39, 0x86, 0xB7, 0x5F, 0x0A, 0x90, 0x6D, 0xEA, 0xF5, 0x68, 0x14, 0x7A, 0x45, 0xA6, 0xDC, 0x87, 0x96, 0x2A, 0x78, 0xA7, 0x3B, 0x60, 0x1C, 0x06, 0xC6, 0x7F, 0xFA, 0x53, 0x83, 0x51, 0x7E, 0xD6, 0x18, 0x42, 0x3C, 0x3C, 0x50, 0xD9, 0x57, 0x7B, 0x10, 0xA5, 0x4E, 0x22, 0xC6, 0x19, 0xB2, 0x4A, 0xC0, 0x94, 0xB8, 0xCF, 0x6F, 0xD5, 0x02, 0xCE, 0x97, 0x16, 0x5E, 0x85, 0x3B, 0xA9, 0x02, 0x3F, 0xDB, 0xE1, 0x3D, 0xF0, 0xBD, 0xCF, 0x22, 0x90, 0x5D, 0x03, 0x42, 0x3C, 0x47, 0x88, 0x2E, 0xD5, 0x18, 0xA4, 0xF9, 0xD0, 0x3B, 0xBB, 0x49, 0x61, 0x70, 0x01, 0x17, 0x6C, 0xC4, 0xB0, 0xD6, 0x75, 0xEA, 0x1B, 0x37, 0x49, 0xF8, 0xED, 0x7F, 0xA9, 0x26, 0xCC, 0x06, 0xE4, 0xE4, 0x65, 0x68, 0x05, 0xDC, 0x52, 0x0B, 0x71, 0x5E, 0x08, 0xC2, 0x79, 0xCF, 0x59, 0x26, 0xF5]

    str = "80 01 2B 00 80 10 2B 00 80 20 2B 00 80 28 2B 00 92 85 2B 00 92 86 2B 00 80 21 2B 00 80 29 2B 00 92 04 2B 00 92 05 2B 00 80 73 2B 00 80 75 2B 00 80 04 2B 00 92 41 2B 00 80 13 2B 00 92 42 2B 00 81 00 2B 00 80 06 2B 00 80 30 2B 00 80 38 2B 00 80 01 11 00 80 01 12 00 80 01 1A 40 80 01 1B 40 92 2D 95 8B 80 10 11 00 80 10 12 00 80 10 1A 40 80 10 1B 40 92 2E 95 8B 80 20 11 00 80 20 12 00 80 20 1A 40 80 20 1B 40 80 20 95 8B 80 28 11 00 80 28 12 00 80 28 1A 40 80 28 1B 40 80 28 95 8B 92 85 11 00 92 85 12 00 92 85 1A 40 92 85 1B 40 92 85 95 8B 92 86 11 00 92 86 12 00 92 86 1A 40 92 86 1B 40 92 86 95 8B 80 21 11 00 80 21 12 00 80 21 1A 40 80 21 1B 40 80 21 95 8B 80 29 11 00 80 29 12 00 80 29 1A 40 80 29 1B 40 80 29 95 8B 92 04 11 00 92 04 12 00 92 04 1A 40 92 04 1B 40 92 25 95 8B 92 05 11 00 92 05 12 00 92 05 1A 40 92 05 1B 40 92 26 95 8B 80 73 11 00 80 73 12 00 80 73 1A 40 80 73 1B 40 92 27 95 8B 80 75 11 00 80 75 12 00 80 75 1A 40 80 75 1B 40 80 75 95 8B 80 04 11 00 92 41 11 00 80 13 11 00 92 42 11 00 81 00 11 00 80 06 11 00 80 30 11 00 80 38 11 00 80 04 12 00 92 41 12 00 80 13 12 00 92 42 12 00 81 00 12 00 80 06 12 00 80 30 12 00 80 38 12 00 80 04 1A 40 92 41 1A 40 80 13 1A 40 92 42 1A 40 81 00 1A 40 80 06 1A 40 80 30 1A 40 80 38 1A 40 80 04 1B 40 92 41 1B 40 80 13 1B 40 92 42 1B 40 81 00 1B 40 80 06 1B 40 80 30 1B 40 80 38 1B 40 80 04 95 8B 92 41 95 8B 80 13 95 8B 92 42 95 8B 81 00 95 8B 80 06 95 8B 80 30 95 8B 80 38 95 8B 80 DF 11 00 80 DF 12 00 80 DF 13 00 80 DF 1E 00 80 DF 95 00 92 33 95 00 92 53 95 00 92 55 95 00 92 54 95 00 92 56 95 00 92 57 95 00 92 33 12 00 92 34 12 00 92 36 12 00 92 35 12 00 80 C7 12 00 92 56 12 00 92 57 12 00 80 C7 95 00 80 98 11 40 80 97 11 40 80 90 11 00 80 98 12 40 80 97 12 40 80 90 12 00 80 98 13 40 80 97 13 40 80 90 13 00 80 98 95 40 80 97 95 40 80 90 95 00 80 98 00 40 80 97 00 40 80 90 00 00 80 98 96 40 80 97 96 40 80 90 96 00 80 95 11 00 80 95 12 00 80 95 13 00 80 95 95 00 80 95 00 00 80 95 96 00 80 91 11 00 80 91 12 00 80 91 13 00 80 91 95 00 80 91 00 00 80 91 96 00 92 87 54 00 92 87 99 00 80 96 11 00 80 96 12 00 80 96 13 00 80 96 95 00 80 96 00 00 80 96 96 00 92 50 16 00 92 7F 47 00 92 50 17 00 92 19 47 40 92 51 49 00 80 DE 09 00 92 84 13 00 92 84 12 00 92 84 11 00 92 2F 55 00 C0 73 88 00 D3 00 55 00 F0 00 51 00 92 51 57 00 92 2C 00 40 C1 00 87 00 C1 29 87 00 C1 40 87 00 C1 55 87 00 92 40 00 00 92 16 47 40 92 18 47 40 92 97 47 40 92 98 47 40 C2 93 87 40"
    bbb = zd_utility.ConvertStr2Bytes(str)
    filter_ = {}
    cnt_ = 0
    for i in range(0, len(bbb), 4):
        type_ = zd_utility.ConvertBytes2Hex([bbb[i+2]])
        dtc_ = zd_utility.ConvertBytes2Hex(bbb[i+0:i+2])
        if type_ in filter_:
            filter_[type_].append(dtc_)
        else:
            list_ = []
            list_.append(dtc_)
            filter_[type_] = list_
        cnt_ += 1
    print(cnt_)
    for key_, value_ in filter_.items():
        print(key_)
        for v in value_:
            print('    '+v)

